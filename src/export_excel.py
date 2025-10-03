import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

def export_to_excel(dataframes_dict, filename):
    os.makedirs("exports", exist_ok=True)
    filepath = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filepath)
    for sheet_name in dataframes_dict.keys():
        ws = wb[sheet_name]

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            if all(cell.value is None or isinstance(cell.value, (int, float)) for cell in col):
                col_letter = col[0].column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    wb.save(filepath)

    total_rows = sum(len(df) for df in dataframes_dict.values())
    print(f"Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")